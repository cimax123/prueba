import streamlit as st
import pandas as pd
import openpyxl
import io
import re
import zipfile
import xml.etree.ElementTree as ET
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Extractor Definitivo V18", layout="wide")

# --- FUNCIONES DE LIMPIEZA ---
def clean_text(text):
    if text:
        return str(text).strip()
    return ""

def normalize_key(text):
    if text:
        return str(text).upper().strip().replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U')
    return ""

def get_float(s):
    if s is None: return 0.0
    try:
        s_clean = re.sub(r'[^\d.,-]', '', str(s)).replace(',', '')
        return float(s_clean)
    except:
        return 0.0

# --- MOTOR 1: LECTURA ESTRUCTURADA (CELDAS) ---
def extract_cell_data(sheet):
    """Extrae datos basados en la grilla de celdas tradicional."""
    data = {}
    
    # Mapeo de celdas para acceso rápido
    cells = {}
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.value:
                cells[(cell.row, cell.column)] = str(cell.value)

    # Función auxiliar de búsqueda vecina
    def find_near(keywords, look_down=True):
        found_val = None
        for (r, c), val in cells.items():
            val_norm = normalize_key(val)
            if any(normalize_key(k) in val_norm for k in keywords):
                # Encontró etiqueta, buscar vecino
                # 1. Derecha
                if (r, c+1) in cells: return cells[(r, c+1)]
                if (r, c+2) in cells: return cells[(r, c+2)] # Salto de 1
                # 2. Abajo
                if look_down:
                    if (r+1, c) in cells: return cells[(r+1, c)]
                    if (r+2, c) in cells: return cells[(r+2, c)]
        return None

    data['Cliente'] = find_near(['CLIENTE', 'CUSTOMER', 'SOLD TO'])
    data['Num_Exp'] = find_near(['EXP', 'INVOICE NO', 'FACTURA'])
    
    # Fecha (Lógica simple)
    data['Fecha'] = find_near(['FECHA', 'DATE'])
    if not data['Fecha']:
        # Intentar buscar año/mes/dia sueltos
        y = find_near(['YEAR', 'AÑO'])
        m = find_near(['MONTH', 'MES'])
        d = find_near(['DAY', 'DIA'])
        if y and m and d: data['Fecha'] = f"{d}-{m}-{y}"

    data['Puerto_Emb'] = find_near(['PUERTO DE EMBARQUE', 'LOADING PORT'])
    data['Puerto_Dest'] = find_near(['PUERTO DESTINO', 'DESTINATION'])
    data['Forma_Pago'] = find_near(['FORMA DE PAGO', 'PAYMENT TERMS'])
    
    # Moneda e Incoterm (Búsqueda global simple)
    all_text = " ".join(cells.values()).upper()
    
    # Moneda
    if "DOLAR" in all_text or "USD" in all_text: data['Moneda'] = "USD"
    elif "EURO" in all_text or "EUR" in all_text: data['Moneda'] = "EUR"
    else: data['Moneda'] = None
    
    # Incoterm
    incos = ['FOB', 'CIF', 'CFR', 'EXW', 'FCA']
    data['Incoterm'] = next((i for i in incos if i in all_text), None)

    return data

# --- MOTOR 2: EXTRACCIÓN DE PRODUCTOS ---
def extract_products(sheet):
    products = []
    
    # Buscar fila de cabecera
    header_row = None
    cols = {}
    
    for row in sheet.iter_rows():
        txts = [normalize_key(c.value) for c in row]
        if any("DESCRIPCION" in t for t in txts) and any("CANTIDAD" in t for t in txts):
            header_row = row[0].row
            for c in row:
                v = normalize_key(c.value)
                if "DESCRIPCION" in v: cols['desc'] = c.column
                elif "CANTIDAD" in v: cols['qty'] = c.column
                elif "PRECIO" in v: cols['price'] = c.column
                elif "TOTAL" in v and "CASES" not in v: cols['total'] = c.column
            break
    
    if header_row and 'desc' in cols:
        for i in range(header_row + 1, sheet.max_row + 1):
            desc = sheet.cell(row=i, column=cols['desc']).value
            desc_str = clean_text(desc)
            
            # Stop condition
            if "TOTAL" in desc_str.upper() and len(desc_str) < 20: break
            
            qty = get_float(sheet.cell(row=i, column=cols['qty']).value if 'qty' in cols else 0)
            price = get_float(sheet.cell(row=i, column=cols['price']).value if 'price' in cols else 0)
            total = get_float(sheet.cell(row=i, column=cols['total']).value if 'total' in cols else 0)
            
            if price > 0:
                products.append({
                    'Descripcion': desc_str,
                    'Cantidad': qty,
                    'Precio Unitario': price,
                    'Total Linea': total if total > 0 else qty*price
                })
    return products

# --- MOTOR 3: "RAW XML SCANNER" (Opción Nuclear) ---
def extract_all_xml_strings(uploaded_file):
    """
    Abre el XLSX como ZIP y extrae TODO el texto plano encontrado
    en SharedStrings y Drawings. Ignora estructura de celdas.
    """
    found_strings = set()
    try:
        uploaded_file.seek(0)
        with zipfile.ZipFile(uploaded_file, 'r') as z:
            # 1. Shared Strings (Texto de celdas)
            if 'xl/sharedStrings.xml' in z.namelist():
                with z.open('xl/sharedStrings.xml') as f:
                    tree = ET.parse(f)
                    for t in tree.iter():
                        if t.text and t.text.strip():
                            found_strings.add(t.text.strip())
            
            # 2. Drawings / Shapes (Texto flotante oculto)
            drawing_files = [x for x in z.namelist() if 'drawings/drawing' in x]
            for df in drawing_files:
                with z.open(df) as f:
                    content = f.read().decode('utf-8', errors='ignore')
                    # Regex sucio para sacar texto entre tags XML en drawings
                    # Las etiquetas suelen ser <a:t>Texto</a:t>
                    matches = re.findall(r'>([^<]+)<', content)
                    for m in matches:
                        if len(m.strip()) > 1: # Ignorar caracteres sueltos
                            found_strings.add(m.strip())
                            
            # 3. Comments
            comment_files = [x for x in z.namelist() if 'comments' in x]
            for cf in comment_files:
                 with z.open(cf) as f:
                    content = f.read().decode('utf-8', errors='ignore')
                    matches = re.findall(r'>([^<]+)<', content)
                    for m in matches:
                        if len(m.strip()) > 1:
                            found_strings.add(m.strip())
                            
    except Exception as e:
        st.error(f"Error XML: {e}")
    
    return list(found_strings)

# --- PROCESAMIENTO PRINCIPAL ---
def process_full(uploaded_file):
    # 1. Leer estructura normal (Openpyxl)
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet = wb.active
        cell_data = extract_cell_data(sheet)
        prods = extract_products(sheet)
    except Exception as e:
        return None, f"Error archivo: {e}", []

    # 2. Leer "Sopa de Letras" (XML Raw)
    raw_texts = extract_all_xml_strings(uploaded_file)
    
    # 3. Búsqueda de "Datos Fantasma" en la Sopa de Letras
    # Observaciones: Buscamos el texto más largo que no sea basura
    obs_candidates = [t for t in raw_texts if len(t) > 30]
    # Filtrar basura común
    obs_clean = []
    for t in obs_candidates:
        tn = normalize_key(t)
        if "SOCIEDAD" in tn or "COMMERCIAL INVOICE" in tn or "BANK" in tn: continue
        obs_clean.append(t)
    
    # Si encontramos algo largo que no es cabecera, probablemente es la observación
    cell_data['Observaciones_Detectadas'] = obs_clean[0] if obs_clean else ""

    # Condición de Venta: Buscamos frases mágicas en el texto crudo
    condicion_candidata = None
    for t in raw_texts:
        tn = normalize_key(t)
        if "BAJO CONDICION" in tn or "CONSIGNACION" in tn or "UNDER CONDITION" in tn:
            condicion_candidata = t
            break
    
    if condicion_candidata:
        cell_data['Condicion_Venta'] = condicion_candidata
    else:
        # Fallback a lo que encontró la celda (aunque sea "COLLECT") si no halló la frase mágica
        # Pero tú dijiste que "COLLECT" está mal. Así que si no encuentra la mágica, lo dejamos vacío
        # o sugerimos al usuario.
        if not cell_data.get('Condicion_Venta'):
             # Intentar buscar la etiqueta en celdas como último recurso
             pass 

    # Tipo de Cambio
    tc_candidate = None
    for t in raw_texts:
        tn = normalize_key(t)
        if "TIPO CAMBIO" in tn or "EXCHANGE RATE" in tn:
            # Intentar extraer numero cercano? Es difícil en raw text.
            tc_candidate = t
            break
    cell_data['Tipo_Cambio_Detectado'] = tc_candidate

    # 4. Unir
    final_rows = []
    base_row = {
        'Archivo': uploaded_file.name,
        'Cliente': cell_data.get('Cliente'),
        'Num_Exp': cell_data.get('Num_Exp'),
        'Fecha': cell_data.get('Fecha'),
        'Condicion_Venta': cell_data.get('Condicion_Venta', ''), # Prioridad Raw
        'Moneda': cell_data.get('Moneda'),
        'Incoterm': cell_data.get('Incoterm'),
        'Observaciones': cell_data.get('Observaciones_Detectadas'),
        'Tipo_Cambio': cell_data.get('Tipo_Cambio_Detectado')
    }
    
    if prods:
        for p in prods:
            row = base_row.copy()
            row.update(p)
            final_rows.append(row)
    else:
        final_rows.append(base_row)
        
    return final_rows, None, raw_texts


# --- UI STREAMLIT ---
st.title("🕵️ Extractor V18: Análisis Forense + Edición Manual")
st.markdown("""
Esta herramienta intenta leer datos ocultos descomprimiendo el Excel.
**Si el dato sigue sin aparecer en la tabla de abajo, es una IMAGEN y no se puede leer.**
Para esos casos, usa la **Tabla Editable** al final para corregir los datos manualmente antes de descargar.
""")

uploaded_files = st.file_uploader("Sube tus Excel", type=['xlsx'], accept_multiple_files=True)

if uploaded_files:
    all_data = []
    raw_debug = {}
    
    with st.spinner("Descomprimiendo archivos y analizando ADN..."):
        for f in uploaded_files:
            rows, err, raw = process_full(f)
            if rows:
                all_data.extend(rows)
                raw_debug[f.name] = raw
            if err:
                st.error(f"{f.name}: {err}")

    if all_data:
        df = pd.DataFrame(all_data)
        
        # Reordenar columnas
        cols = ['Archivo', 'Cliente', 'Num_Exp', 'Fecha', 'Condicion_Venta', 'Incoterm', 
                'Moneda', 'Tipo_Cambio', 'Cantidad', 'Descripcion', 'Precio Unitario', 
                'Total Linea', 'Observaciones']
        final_cols = [c for c in cols if c in df.columns] + [c for c in df.columns if c not in cols]
        df = df[final_cols]

        st.divider()
        st.subheader("📝 Tabla de Resultados Editable")
        st.info("Haz doble clic en cualquier celda para corregir datos faltantes (ej: Condición de Venta, Observaciones).")

        # --- AG GRID (TABLA EDITABLE) ---
        gb = GridOptionsBuilder.from_dataframe(df)
        gb.configure_default_column(editable=True, resizable=True, filterable=True, sortable=True)
        gb.configure_column("Archivo", editable=False) # No editar nombre archivo
        gb.configure_selection('multiple', use_checkbox=True)
        gridOptions = gb.build()

        grid_response = AgGrid(
            df,
            gridOptions=gridOptions,
            update_mode=GridUpdateMode.VALUE_CHANGED, # Actualizar al editar
            fit_columns_on_grid_load=False,
            height=400,
            theme='streamlit'
        )

        df_edited = grid_response['data']
        
        # Botón descarga datos corregidos
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            pd.DataFrame(df_edited).to_excel(writer, index=False)
            
        st.download_button(
            "💾 Descargar Excel Corregido",
            data=buffer.getvalue(),
            file_name="facturas_finales.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

        # --- SECCIÓN DE DIAGNÓSTICO ---
        with st.expander("🔍 Ver todo el texto crudo encontrado (Sopa de Letras)"):
            st.write("Si tu texto 'Bajo Condición' o 'Estas cajas...' NO está en esta lista, es una imagen.")
            for fname, texts in raw_debug.items():
                st.write(f"**{fname}** ({len(texts)} fragmentos de texto encontrados):")
                st.code("\n".join(texts))
